# -*- coding: utf-8 -*-

import traceback
import numfmt
import xlsconfig
import util
from util import ExcelToCodeException, log_error
from tps import tp0, convention

def format_number_with_xlrd(f, cell, wb):
	xf = wb.xf_list[cell.xf_index]
	fmt_key = xf.format_key
	fmt = wb.format_map[fmt_key]
	s_fmt = fmt.format_str
	a_fmt = numfmt.extract_number_format(s_fmt)
	if a_fmt:
		s_f = numfmt.format_number(f, a_fmt, ',', '.')
	else:
		s_f = "%g" % f
	return s_f

def format_number_with_openpyxl(f, cell, wb):
	s_fmt = cell.number_format
	a_fmt = numfmt.extract_number_format(s_fmt)
	if a_fmt:
		return numfmt.format_number(f, a_fmt, ',', '.')
	else:
		return "%g" % f

format_number = format_number_with_openpyxl

class ConverterInfo:
	def __init__(self, info, column = None):
		self.header = info[0]
		self.field = info[1]
		self.convert = info[2]
		self.can_default = info[3] if len(info) > 3 else False
		self.exist_default_value = len(info) > 4
		self.default_value = info[4] if len(info) > 4 else None
		self.column = column

class BaseParser(object):

	def __init__(self, filename, module, sheet_index=0):
		super(BaseParser, self).__init__()

		# 表格key名。调试用
		self.key_name = "ID"
		self.is_multi_key = False

		self.filename = filename
		self.module = module
		self.sheet_index = sheet_index

		# 转换器。列索引 -> 转换器(ConverterInfo)
		self.converters = {}

		# 最终生成的数据表。key是第一格对应的数据。
		# 如果表格是multi_key，则value是一个数组，包含了所有key相同的行。
		self.sheet = {}

		# 表头的类型数据。
		self.sheet_types = {}

		# 表头映射到列索引
		self.header_2_col = {}
		# 表头数组。也就是连续的表头列数，空列右侧的内容会被忽略。
		self.headers = []

		# 每一列的默认值。列索引 -> 默认值
		self.default_values = {}

		self.argument_row_index = xlsconfig.SHEET_ROW_INDEX["argument"]
		self.header_row_index 	= xlsconfig.SHEET_ROW_INDEX["header"]
		self.data_row_index 	= xlsconfig.SHEET_ROW_INDEX["data"]
		self.default_value_row_index = xlsconfig.SHEET_ROW_INDEX.get("default", -1)

		# 表格是否是垂直排列
		self.is_vertical = False
		# 如果是垂直排列的话，从第几行开始才是真正的数据
		self.vertical_start_row = xlsconfig.SHEET_ROW_INDEX.get("verticalStartRow", 2)

		self.workbook = None
		self.worksheet = None

		self.version = None

		# excel表头第一行的参数
		self.arguments = {}

		# 上一次解析行的key
		self.last_key = None

	def run(self):
		self.do_parse()

	# 将单元格数据转换成字符串形式。
	# 因为'123,456'格式的字符串，会被Excel存贮成浮点数：123456，
	# 而openpyxl仅仅是读取存贮的数据，不会自动将数字还原成字符串，所以这里手动进行转换。
	def extract_cell_value(self, cell):
		value = cell.value
		tp = type(value)

		if value is None:
			value = ""
		elif tp == float or tp == long or tp == int:
			value = format_number(value, cell, self.workbook)
		elif tp == unicode:
			value = value.encode("utf-8").strip()
		elif tp == str:
			value = value.strip()
		else:
			msg = "不支持的数据类型: %s -> '%s'" % (type(value), value)
			raise ExcelToCodeException, msg

		return value

	def get_default_value(self, col):
		value = self.default_values.get(col, "")
		if value == '!':
			raise ExcelToCodeException, "该项必填"
		value = value.replace("\\!", "!")
		return value

	def convert_cell(self, row, col, value):
		converter = self.converters[col]
		if converter is None:
			return None

		if value == "":
			value = self.get_default_value(col)

		ret = None
		if value == "":
			if not converter.can_default:
				raise ExcelToCodeException, "该项必填"

		else:
			try:
				ret = converter.convert(value)
			except:
				raise ExcelToCodeException, "类型转换失败(%s)" % (str(converter.convert), )

		if ret is None and converter.exist_default_value:
			ret = converter.default_value

		return ret

	def do_parse(self):
		import openpyxl
		self.workbook = openpyxl.load_workbook(self.filename)

		sheets = self.workbook.worksheets
		if self.sheet_index >= len(sheets):
			log_error("Excel表'%s'没有子表'%d'", self.filename, self.sheet_index)
			return

		worksheet = sheets[self.sheet_index]
		self.worksheet = worksheet

		# 注意：worksheet的行列索引是从"1"开始的

		if worksheet.max_column > 1000:
			print "warn: Excel '%s' 列数过多：%d" % (self.filename, worksheet.max_column)
		if worksheet.max_row > 32767:
			print "warn: Excel '%s' 行数过多：%d" % (self.filename, worksheet.max_row)

		self.parse_arguments(self.extract_cells(self.argument_row_index))

		if self.version is None:
			log_error("无效的数据表文件，必须存在版本号信息")
			return

		self.parse_header(self.extract_cells(self.header_row_index))
		if self.default_value_row_index >= 0:
			self.parse_defaults(self.extract_cells(self.default_value_row_index))

		if self.is_vertical:
			self.parse_by_vertical(worksheet)
		else:
			self.parse_by_horizontal(worksheet)
		return

	def is_blank_line(self, cells, count):
		for i in enumerate(count):
			v = cells[i].value
			if v != '' and v is not None:
				return False

		return True

	def parse_cells(self, r, cells):
		max_cols = len(self.headers)

		# 遇到空白行，表示解析完成
		if self.is_blank_line(cells, max_cols):
			return

		# 如果key是空，自动复制上一行key
		first_value = cells[0].value
		if first_value == '':
			cells[0].value = self.last_key
		else:
			self.last_key = first_value

		current_row_data = []
		for c in xrange(max_cols):
			cell_value = None
			result_value = None
			try:
				cell_value = self.extract_cell_value(cells[c])
				result_value = self.convert_cell(r, c, cell_value)
			except ExcelToCodeException, e:
				# traceback.print_exc()
				log_error("单元格 %s = [%s] 数据解析失败。原因：%s", self.row_col_str(r, c), str(cell_value), str(e))

			current_row_data.append(result_value)

		self.add_row(current_row_data)
		return True

	def parse_by_horizontal(self, worksheet):
		for r in xrange(self.data_row_index, worksheet.max_row):
			cells = worksheet[r + 1]
			if not self.parse_cells(r, cells):
				break
		return

	def parse_by_vertical(self, worksheet):
		for c in xrange(self.data_row_index, worksheet.max_column):
			col_index = util.int_to_base26(c)
			cells = worksheet[col_index][self.vertical_start_row : ]
			if not self.parse_cells(c, cells):
				break
		return

	def add_row(self, current_row_data):
		key_value = current_row_data[0]
		
		if self.is_multi_key:
			row = self.sheet.setdefault(key_value, [])
			row.append(current_row_data)

		else:
			if key_value in self.sheet:
				raise ExcelToCodeException, "Key'%s'重复" % key_value

			self.sheet[key_value] = current_row_data

	def parse_header(self, cells):
		self.headers = cells

		for col, header in enumerate(cells):
			if header == "":
				self.headers = cells[:col]
				break

			if header in self.header_2_col:
				log_error("表头'%s'重复，at %s", header, self.row_col_str(self.header_row_index, col))
				continue

			self.header_2_col[header] = col

		if len(self.headers) == 0:
			return log_error("Except表'%s'表头数量不能为0", self.filename)

		return

	def parse_arguments(self, cells):
		compatible_posfix = '：'

		self.arguments = {}
		for col in xrange(0, len(cells), 2):
			header = cells[col]
			if header is None: break

			if header.endswith(compatible_posfix):
				split = len(header) - len(compatible_posfix)
				header = header[:split]

			converter = xlsconfig.ARGUMENT_CONVERTER.get(header)
			if converter is None: continue

			field, type = converter
			method = convention.type2function(type)

			value = cells[col + 1]
			ret = None
			try:
				ret = method(value)
			except:
				traceback.print_exc()

				log_error("参数转换失败，%s = [%s]", self.row_col_str(self.argument_row_index, col), value)

			self.arguments[field] = ret

		self.version = self.arguments.get("version")

		self.is_vertical = self.arguments.get("vertical", False)
		if self.is_vertical:
			self.header_row_index -= self.vertical_start_row
			self.data_row_index -= self.vertical_start_row
			self.default_value_row_index -= self.vertical_start_row

		multi_key = self.arguments.get("multiKey")
		if multi_key is not None:
			self.is_multi_key = multi_key
		return

	def parse_defaults(self, cells):
		default_values = {}
		for col, value in enumerate(cells):
			if value != "":
				default_values[col] = value

		self.default_values = default_values

	def extract_cells(self, index):
		if self.is_vertical:
			col = util.int_to_base26(index)

			ret = []
			columns = self.worksheet[col]
			for i in xrange(self.vertical_start_row, len(columns)):
				ret.append(self.extract_cell_value(columns[i]))

			return ret
		else:
			row = index + 1
			return [self.extract_cell_value(cell) for cell in self.worksheet[row]]

	def row_col_str(self, row, col):
		real_r, real_c = row, col
		if self.is_vertical:
			real_r = col + self.vertical_start_row
			real_c = row
		return "%d:%s" % (real_r + 1, util.int_to_base26(real_c))
