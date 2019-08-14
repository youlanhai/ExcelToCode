# -*- coding: utf-8 -*-
"""
将Excel表转换成csv

Excel的解析速度比较慢，可以用多进程转换为csv。
然后再启动导表工具，进行导出处理。

也可以用自定义的Excel-CSV转换工具处理完表格，比如Excel表是分页格式。
然后再启动导表工具，进行导出处理。
"""

from os import path
from argparse import ArgumentParser
import csv

import numfmt
import util
from util import _S, log_error, ExcelToCodeException

def main():
	parser = ArgumentParser(description = _S("将excel表格转换成csv文件"))
	parser.add_argument("-o", "--output", help=_S("输出文件路径。如果输入路径是目录，输出路径也应该是目录"))
	parser.add_argument("-a", "--all-worksheet", action="store_true", help=_S("是否导出所有的sheet页"))
	parser.add_argument("input", help=_S("输入文件路径。可以是文件，也可以是目录"))

	option = parser.parse_args()
	input_path = option.input
	output_path = option.output

	inputs = None
	if path.isfile(input_path):
		if output_path is None:
			output_path = path.splitext(input_path)[0] + ".csv"
		xl2csv(input_path, output_path, option)
	else:
		if output_path is None:
			output_path = input_path

		files = util.gather_all_files(input_path, (".xlsx", ))
		for fname in files:
			src_path = path.join(input_path, fname)
			dst_path = path.join(output_path, path.splitext(fname)[0] + ".csv")
			xl2csv(src_path, dst_path, option)

	if util.has_error:
		exit(-1)
	return


def xl2csv(input_path, output_path, option):
	print "parse:", input_path
	parser = Parser(
		input_path,
		export_all_sheet = option.all_worksheet
	)
	parser.parse()
	parser.save(output_path)


# 因为'123,456'格式的字符串，会被Excel存贮成浮点数：123456，
# 而openpyxl仅仅是读取存贮的数据，不会自动将数字还原成字符串，所以这里手动进行转换。
def format_number(f, cell, parser):
	a_fmt = numfmt.extract_number_format(cell.number_format)
	if a_fmt:
		return numfmt.format_number(f, a_fmt, ',', '.')
	else:
		s = "%f" % f
		# 删除'.'后面的0
		return remove_end_zero(s)

FORMAT_MAP = {
	type(None) : lambda value, cell, parser: "",
	int 	: format_number,
	long 	: format_number,
	float 	: format_number,
	bool 	: lambda value, cell, parser: str(value),
	str 	: lambda value, cell, parser: value.strip(),
	unicode : lambda value, cell, parser: value.encode("utf-8").strip()
}

class Parser(object):

	FORMAT_MAP = FORMAT_MAP

	def __init__(
		self,
		filename,
		sheet_index = 0,
		export_all_sheet = False,
		max_row = 65535,
		max_column = 1000,
	):
		super(Parser, self).__init__()

		# 表格路径
		self.filename = filename

		self.sheet = []
		self.sheets = []
		self.sheet_names = []

		self.sheet_index = 0
		self.export_all_sheet = export_all_sheet

		self.max_row = max_row
		self.max_column = max_column

		# 当前的Excel表
		self.workbook = None
		# 当前的sheet页
		self.worksheet = None

	# 将单元格数据转换成字符串形式。
	def extract_cell_value(self, cell):
		value = cell.value
		tp = type(value)

		method = self.FORMAT_MAP.get(tp)
		if method:
			return method(value, cell, self)

		raise ExcelToCodeException, "不支持的数据类型: %s" % str(tp)

	def parse(self):
		try:
			import openpyxl
		except:
			return log_error("请安装插件: openpyxl")

		self.workbook = openpyxl.load_workbook(self.filename.decode("utf-8"))

		worksheets = self.workbook.worksheets
		self.sheets = [None] * len(worksheets)
		self.sheet_names = [None] * len(worksheets)

		if self.export_all_sheet:
			for i, worksheet in enumerate(worksheets):
				self.parse_sheet(i, worksheet)
		else:
			if self.sheet_index >= len(worksheets):
				log_error("Excel表'%s'没有子表'%d'", self.filename, self.sheet_index)
			else:
				self.parse_sheet(self.sheet_index, worksheets[self.sheet_index])
		return

	def parse_sheet(self, i, worksheet):
		self.worksheet = worksheet
		# print "all methods:", dir(worksheet)

		# 注意：worksheet的行列索引是从"1"开始的

		if worksheet.max_column > self.max_column:
			print _S("warn: Excel '%s' 列数过多：%d" % (self.filename, worksheet.max_column))
		if worksheet.max_row > self.max_row:
			print _S("warn: Excel '%s' 行数过多：%d" % (self.filename, worksheet.max_row))

		self.current_sheet_index = i
		self.sheet = []
		self.sheets[i] = self.sheet
		self.sheet_names[i] = worksheet.title

		row_count = min(worksheet.max_row, self.max_row)
		col_count = min(worksheet.max_column, self.max_column)

		for r in xrange(row_count):
			cells = worksheet[r + 1]

			# 遇到空白行，表示解析完成
			if self.is_blank_line(cells, col_count):
				break

			if not self.parse_cells(r, cells, col_count):
				break

		return

	def is_blank_line(self, cells, col_count):
		for i in xrange(col_count):
			v = cells[i].value
			if v != '' and v is not None:
				return False

		return True

	def parse_cells(self, r, cells, col_count):
		current_row_data = []
		for c in xrange(col_count):
			value = None
			try:
				value = self.extract_cell_value(cells[c])
			except ExcelToCodeException, e:
				value = str(cells[c].value)
				log_error("单元格 %s = [%s] 数据解析失败。原因：%s", self.row_col_str(r, c), value, str(e))

			current_row_data.append(value)

		self.sheet.append(current_row_data)
		return True

	def row_col_str(self, row, col):
		return "%d:%s" % (row, util.int_to_base26(col))

	def save(self, output_path):
		if self.export_all_sheet:
			output_name = path.splitext(output_path)[0]
			for i, sheet in enumerate(self.sheets):
				sheet_name = self.sheet_names[i]
				sheet_path = "%s_%s.csv" % (output_name, sheet_name)
				self.save_sheet(sheet, sheet_path)
		else:
			self.save_sheet(self.sheet, output_path)

	def save_sheet(self, sheet, output_path):
		util.ensure_folder_exist(output_path)
		with open(output_path, "wb") as f:
			writer = csv.writer(f)
			writer.writerows(sheet)

def remove_end_zero(s):
	pos = s.find('.')
	if pos < 0: return s
	
	i = len(s)
	while i > pos and s[i - 1] == '0':
		i -= 1
	if s[i - 1] == '.':
		i -= 1
	return s[:i]

if __name__ == "__main__":
	main()
