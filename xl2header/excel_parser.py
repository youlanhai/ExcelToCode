# -*- coding: utf-8 -*-
import json
from os import path
import openpyxl

from . import xlsconfig
from . import util
from .util import ExcelToCodeException, row_col_str, log
from . import header_util
from .header import Header, Argument

def format_number(f, cell, parser):
	return str(f)


FORMAT_MAP = {
	type(None) : lambda value, cell, parser: "",
	int 	: format_number,
	float 	: format_number,
	bool 	: lambda value, cell, parser: str(value).upper(),
	str 	: lambda value, cell, parser: value.strip(),
}


class ExcelParser(object):

	FORMAT_MAP = FORMAT_MAP

	def __init__(
		self,
		filename,
		sheet_index = 0,
		force = False,
		auto_create = False
	):
		super(ExcelParser, self).__init__()

		# 表格路径
		self.filename = filename

		self.sheet_index = sheet_index

		# 当前的Excel表
		self.workbook = None
		# 当前的sheet页
		self.worksheet = None

		self.header_root = Header.new_root()
		self.arguments = Argument()

		self.argument_row_index = xlsconfig.SHEET_ROW_INDEX["argument"]

		# 表格是否是垂直排列
		self.is_vertical = False

		self.force = force
		self.auto_create = auto_create

	# 将单元格数据转换成字符串形式。
	def parse_cell_value(self, cell):
		value = cell.value
		tp = type(value)

		method = self.FORMAT_MAP.get(tp)
		if method:
			return method(value, cell, self)

		addr = self.sheet_proxy.get_cell_addr(cell)
		raise ExcelToCodeException("不支持的数据类型: %s, at %s" % (str(tp), addr), self.filename)

	def parse(self):
		u_filename = self.filename
		if not path.exists(u_filename):
			if not self.auto_create:
				raise ExcelToCodeException("Excel文件不存在", self.filename)
			return

		self.workbook = openpyxl.load_workbook(u_filename)
		self.worksheet = self.workbook.worksheets[self.sheet_index]

		self.sheet_proxy = HorizontalSheetProxy(self, self.worksheet)

		self.parse_arguments()

		if self.is_vertical:
			self.sheet_proxy = VerticalSheetProxy(self, self.worksheet)

		self.sheet_proxy.init()

		self.parse_header()
		return

	def generate_header(self, header_file):
		new_header, new_arguments = header_util.load_header_list(header_file)

		# 更新索引
		for i, header in enumerate(new_header.children):
			old = self.header_root.find_child(header.title)
			header.index = old.index if old else -1

		if not self.force and not self.is_changed(new_header, new_arguments):
			return

		if not self.workbook:
			vertical = new_arguments.find("垂直排列")
			self.is_vertical = vertical == "TRUE" or vertical == "1"

			log("create excel: %s", self.filename)
			self.create_new_excel()

		log("generate header for: " + self.filename)
		self.sheet_proxy.generate_header(new_header)
		self.sheet_proxy.generate_arguments(new_arguments)

		u_filename = self.filename
		util.ensure_folder_exist(u_filename)
		self.workbook.save(u_filename)

	def create_new_excel(self):
		self.workbook = openpyxl.Workbook()
		self.worksheet = self.workbook.active

		cls = VerticalSheetProxy if self.is_vertical else HorizontalSheetProxy
		self.sheet_proxy = cls(self, self.worksheet)
		self.sheet_proxy.init()

	def is_changed(self, new_header, new_arguments):
		len1 = len(new_header.children)
		len2 = len(self.header_root.children)
		if len1 != len2:
			log("header length not equal: %d != %d", len1, len2)
			return True

		for index, header in enumerate(new_header.children):
			if index != header.index:
				log("header index not equal: %d != %d", index, header.index)
				return True

			old = self.header_root.find_child(header.title)
			if old is None:
				log("old header not exist")
				return True

			if not header.equal(old):
				log(
					"header value not equal: (%s, %s, %s) != (%s, %s, %s)",
					header.title, header.field, header.field_type, old.title, old.field, old.field_type)
				return True

		len1 = len(new_arguments.values)
		len2 = len(self.arguments.values)
		if len1 != len2:
			log("arguments length not equal: %d != %d", len1, len2)
			return True

		for index, pair in enumerate(new_arguments.values):
			old = self.arguments.values[index]
			if pair != old:
				log("arguments value not equal: (%s, %s) != (%s, %s)", pair[0], pair[1], old[0], old[1])
				return True

		return False

	def parse_header(self):
		header_cells = self.sheet_proxy.get_cells(self.sheet_proxy.header_index)
		field_cells = self.sheet_proxy.get_cells(self.sheet_proxy.field_index)
		type_cells = self.sheet_proxy.get_cells(self.sheet_proxy.type_index)

		root = self.header_root

		for index, cell in enumerate(header_cells):
			title = self.parse_cell_value(cell)
			if not title:
				break

			if root.find_child(title):
				msg = "表头'%s'重复，at %s" % (title, self.sheet_proxy.get_cell_addr(cell))
				raise ExcelToCodeException(msg, self.filename)

			header = Header()
			header.index = index
			header.title = title
			header.field = self.parse_cell_value(field_cells[index])
			header.field_type = self.parse_cell_value(type_cells[index])
			root.add_child(header)

		if len(root.children) == 0:
			raise ExcelToCodeException("表头数量不能为0", self.filename)
		return

	def parse_arguments(self):
		compatible_posfix = '：'

		header = None
		row = self.argument_row_index
		for col in range(self.sheet_proxy.max_column):
			cell = self.sheet_proxy.get_cell(row, col)

			if col % 2 == 0:
				if cell is None or cell.value is None:
					break

				header = self.parse_cell_value(cell)
				if not isinstance(header, str):
					msg = "参数key必须是字符串类型。 value = %s, at %s" % (str(cell.value), self.sheet_proxy.get_cell_addr(cell))
					raise ExcelToCodeException(msg, self.filename)

				if header.endswith(compatible_posfix):
					split = len(header) - len(compatible_posfix)
					header = header[:split]
			else:
				# converter = xlsconfig.ARGUMENT_CONVERTER.get(value)
				# if converter is None:
				# 	continue
				# field, type = converter
				value = self.parse_cell_value(cell)
				self.arguments.add(header, value)

		vertical = self.arguments.find("垂直排列")
		self.is_vertical = vertical == "TRUE" or vertical == "1"
		return

	def save(self, output_path):
		util.ensure_folder_exist(output_path)
		header_util.save_header_list(output_path, self.header_root, self.arguments)

	def save_test(self, output_path):
		tree = header_util.gen_header_tree(self.header_root, self.filename)
		list = header_util.gen_header_list(tree)

		tree_data = []
		self.save_as_tree(tree, tree_data)

		list_data = []
		self.save_as_list(list, list_data)

		arguments = self.arguments.values

		data = {
			"arguments" : arguments,
			# "header" : self.headers,
			"tree" : tree_data,
			"list" : list_data,
		}

		util.ensure_folder_exist(output_path)
		with open(output_path, "w", encoding = "utf-8", newline = "\n") as f:
			json.dump(data, f, indent = 4, sort_keys=True, ensure_ascii = False)

	def save_as_tree(self, node, data):
		self.save_tree_children(node.children, data)

	def save_tree_children(self, children, data):
		for child in children:
			v = {
				"title" 	: child.title,
				"field" 	: child.field,
				"field_type" : child.field_type,
				"type" 		: child.type,
				"index" 	: child.index,
				"end_title" : child.end_title,
				"end_index" : child.end_index,
			}

			if child.children is not None:
				children_data = []
				self.save_tree_children(child.children, children_data)
				v["children"] = children_data

			data.append(v)

	def save_as_list(self, node, data):
		for child in node.children:
			data.append("%s, %s, %s" % (child.title, child.field, child.field_type))


# 注意：worksheet的行列索引是从"1"开始的


class SheetProxy(object):
	def __init__(self, parser, worksheet):
		super(SheetProxy, self).__init__()
		self.parser = parser
		self.worksheet = worksheet
		self.max_row = xlsconfig.MAX_ROW
		self.max_column = xlsconfig.MAX_COLUMN
		self.header_index = xlsconfig.SHEET_ROW_INDEX["header"]
		self.field_index = xlsconfig.SHEET_ROW_INDEX["field"]
		self.type_index = xlsconfig.SHEET_ROW_INDEX["type"]
		self.data_index = xlsconfig.SHEET_ROW_INDEX["data"]

	def init(self):
		pass

	# 获取单元格。索引从0开始
	def get_cell(self, row, col):
		return self.worksheet.cell(row + 1, col + 1)

	def get_cell_addr(self, cell):
		return row_col_str(cell.row - 1, cell.column - 1)

	# 获取一行/列有效数据.
	# 如果是垂直排列的表，则获取一列; 否则获取一行.
	# 索引从0开始
	def get_cells(self, index):
		pass

	# 根据表头描述信息，生成表头
	def generate_header(self, new_header):
		pass

	def generate_arguments(self, arguments):
		for c, (key, value) in enumerate(arguments.values):
			self.worksheet.cell(1, c * 2 + 1).value = key
			self.worksheet.cell(1, c * 2 + 2).value = value
		self.worksheet.cell(1, len(arguments.values) * 2 + 1).value = None


class VerticalSheetProxy(SheetProxy):

	def init(self):
		# 如果是垂直排列的话，从第几行开始才是真正的数据
		self.start_row = xlsconfig.SHEET_ROW_INDEX.get("verticalStartRow", 2)

		self.header_index -= self.start_row
		self.field_index -= self.start_row
		self.type_index -= self.start_row
		self.data_index -= self.start_row

		self.max_row, self.max_column = self.calculate_max_row_column()

	def calculate_max_row_column(self):
		max_column = 0
		max_row = 0

		MAX_ROW, MAX_COLUMN = xlsconfig.MAX_COLUMN, xlsconfig.MAX_ROW

		for c in range(0, MAX_COLUMN):
			cell = self.worksheet.cell(self.start_row + 1, c + 1)
			if cell is None or cell.value is None:
				break
			max_column = c + 1
		for r in range(self.start_row, MAX_ROW):
			cell = self.worksheet.cell(r + 1, self.header_index + 1)
			if cell is None or cell.value is None:
				break
			max_row = r + 1

		return max_row, max_column

	# 获取一列有效数据.
	def get_cells(self, col):
		ret = []
		for row in range(self.start_row, self.max_row):
			ret.append(self.get_cell(row, col))
		return ret

	def generate_header(self, new_header):
		start_row = self.start_row

		worksheet = self.worksheet
		_cells = self.worksheet._cells

		# 提取所有的单元格，然后重新分配位置。
		# 该方法速度快，但是不安全。引用了非公有变量，可能插件升级后就无法使用了。
		rows = {}
		for r in range(start_row, self.max_row):
			cells = []
			for c in range(0, self.max_column):
				cell = worksheet.cell(r + 1, c + 1)
				cells.append(cell)
				del _cells[(cell.row, cell.column)]
			rows[r] = cells

		for r, header in enumerate(new_header.children):
			r += start_row
			if header.index >= 0:
				old_r = start_row + header.index
				for c in range(0, self.max_column):
					cell = rows[old_r][c]
					row = r + 1
					column = c + 1

					_cells[row, column] = cell
					cell.row = row
					cell.column = column
			else:
				# 补一行空数据
				for c in range(self.data_index, self.max_column):
					worksheet.cell(r + 1, c + 1).value = None

			worksheet.cell(r + 1, self.header_index + 1).value = header.title
			worksheet.cell(r + 1, self.field_index + 1).value = header.field
			worksheet.cell(r + 1, self.type_index + 1).value = header.field_type

		# 在末尾补上空的行
		last_row = start_row + len(new_header.children)
		for c in range(0, self.max_column):
			worksheet.cell(last_row + 1, c + 1).value = None
		return


class HorizontalSheetProxy(SheetProxy):

	def init(self):
		self.max_row, self.max_column = self.calculate_max_row_column()

	def calculate_max_row_column(self):
		max_column = 0
		max_row = 0

		for c in range(xlsconfig.MAX_COLUMN):
			cell = self.worksheet.cell(self.header_index + 1, c + 1)
			if cell is None or cell.value is None:
				break
			max_column = c + 1
		for r in range(self.data_index, xlsconfig.MAX_ROW):
			cell = self.worksheet.cell(r + 1, 1)
			if cell is None or cell.value is None:
				break
			max_row = r + 1

		return max_row, max_column

	# 获取一行有效数据.
	def get_cells(self, row):
		ret = []
		for col in range(self.max_column):
			ret.append(self.get_cell(row, col))
		return ret

	def generate_header(self, new_header):
		max_row = self.max_row
		start_row = self.header_index

		worksheet = self.worksheet
		_cells = self.worksheet._cells

		# 提取所有的单元格，然后重新分配位置。
		# 该方法速度快，但是不安全。引用了非公有变量，可能插件升级后就无法使用了。
		rows = {}
		for r in range(start_row, max_row):
			cells = []
			for c in range(self.max_column):
				cell = worksheet.cell(r + 1, c + 1)
				cells.append(cell)
				del _cells[(cell.row, cell.column)]
			rows[r] = cells

		for c, header in enumerate(new_header.children):
			if header.index >= 0:
				for r in range(start_row, max_row):
					cell = rows[r][header.index]
					row = r + 1
					column = c + 1

					_cells[row, column] = cell
					cell.row = row
					cell.column = column
			else:
				# 补一列空数据
				for r in range(self.data_index, max_row):
					worksheet.cell(r + 1, c + 1).value = None

			worksheet.cell(self.header_index + 1, c + 1).value = header.title
			worksheet.cell(self.field_index + 1, c + 1).value = header.field
			worksheet.cell(self.type_index + 1, c + 1).value = header.field_type

		# 在末尾补上空的列
		last_col = len(new_header.children)
		for r in range(start_row, max_row):
			worksheet.cell(r + 1, last_col + 1).value = None
		return

	def generate_header2(self, new_header):
		max_row = self.worksheet.max_row
		start_row = self.header_index
		offset = len(new_header.children) + 1

		# 将原有数据移动到末尾。然后再依次移到相应的位置
		self.worksheet.insert_cols(0, offset)
		for c, header in enumerate(new_header.children):
			if header.index >= 0:
				index = offset + header.index
				col = util.int_to_base26(index)
				range_key = "%s%d:%s%d" % (col, 1, col, max_row + 1)
				self.worksheet.move_range(range_key, cols = c - index)
			else:
				values = [header.title, header.field, header.field_type]
				for r, value in enumerate(values):
					self.worksheet.cell(start_row + r + 1, c + 1).value = value
				for r in range(self.data_index, max_row):
					self.worksheet.cell(r + 1, c + 1).value = None

		# 在末尾补上空的列
		last_col = len(new_header.children)
		for r in range(start_row, max_row):
			self.worksheet.cell(r + 1, last_col + 1).value = None
		return
