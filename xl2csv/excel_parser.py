# -*- coding: utf-8 -*-
from os import path
import csv
import datetime
import time

import util
from util import log, log_error, ExcelToCodeException

FORMAT_MAP = {
	type(None) : lambda value, cell, parser: "",
	int 	: util.format_number,
	float 	: util.format_number,
	bool 	: lambda value, cell, parser: str(value),
	str 	: lambda value, cell, parser: value.strip(),
	datetime.time : lambda value, cell, parser: str(value),
	datetime.datetime : lambda value, cell, parser: str(value),
}

class ExcelParser(object):

	FORMAT_MAP = FORMAT_MAP

	def __init__(
		self,
		filename,
		sheet_index = 0,
		export_all_sheet = False,
		max_row = 65535,
		max_column = 1000,
	):
		super(ExcelParser, self).__init__()

		# 表格路径
		self.filename = filename

		self.sheet = []
		self.sheets = []
		self.sheet_names = []

		self.sheet_index = 0
		self.export_all_sheet = export_all_sheet

		self.max_row = max_row
		self.max_column = max_column

		# 当前的Excel表
		self.workbook = None
		# 当前的sheet页
		self.worksheet = None

	# 将单元格数据转换成字符串形式。
	def parse_cell_value(self, cell):
		value = cell.value
		tp = type(value)
		method = self.FORMAT_MAP.get(tp)
		if method:
			return method(value, cell, self)

		raise ExcelToCodeException("不支持的数据类型: %s" % str(tp))

	def parse(self):
		try:
			import openpyxl
		except:
			return log_error("请安装插件: openpyxl")

		t = time.time()
		# 支持excel公式
		self.workbook = openpyxl.load_workbook(self.filename, data_only = True)
		t = time.time() - t
		if t > 0.5:
			print("加载表格耗时较长:", t, self.filename)

		worksheets = self.workbook.worksheets
		self.sheets = [None] * len(worksheets)
		self.sheet_names = [None] * len(worksheets)

		if self.export_all_sheet:
			for i, worksheet in enumerate(worksheets):
				self.parse_sheet(i, worksheet)
		else:
			if self.sheet_index >= len(worksheets):
				log_error("%s, 没有子表'%d'", self.filename, self.sheet_index)
			else:
				self.parse_sheet(self.sheet_index, worksheets[self.sheet_index])
		return

	def parse_sheet(self, i, worksheet):
		self.worksheet = worksheet

		self.current_sheet_index = i
		self.sheet = []
		self.sheets[i] = self.sheet
		self.sheet_names[i] = worksheet.title

		t = time.time()

		parser = ValidCellsParser(worksheet, filename = self.filename)
		parser.parse()

		self.max_row = parser.max_rows
		self.max_column = parser.max_cols

		t = time.time() - t
		if t > 0.3:
			print("解析数据耗时较长:", t, "rows:", self.max_row, "cols:", self.max_column, self.filename)

		for r, cells in enumerate(parser.sheet):
			if not self.parse_cells(r, cells):
				break

		return

	def parse_cells(self, r, cells):
		current_row_data = []
		for c, cell in enumerate(cells):
			value = None
			try:
				value = self.parse_cell_value(cell)
			except ExcelToCodeException as e:
				value = str(cells[c].value)
				log_error("%s, 单元格 %s = [%s] 数据解析失败。原因：%s", self.filename, self.row_col_str(r, c), value, str(e))

			current_row_data.append(value)

		self.sheet.append(current_row_data)
		return True

	def row_col_str(self, row, col):
		return "%d:%s" % (row, util.int_to_base26(col))

	def save(self, output_path):
		if self.export_all_sheet:
			output_name = path.splitext(output_path)[0]
			for i, sheet in enumerate(self.sheets):
				sheet_name = self.sheet_names[i]
				sheet_path = "%s_%s.csv" % (output_name, sheet_name)
				self.save_sheet(sheet, sheet_path)
		else:
			self.save_sheet(self.sheet, output_path)

	def save_sheet(self, sheet, output_path):
		util.ensure_folder_exist(output_path)
		with open(output_path, "w", encoding = "utf-8", newline='') as f:
			writer = csv.writer(f)
			writer.writerows(sheet)


""" 解析有效的行列数量
算法说明:
1. 横向扫描，当扫描到第c列时，如果值为空，且再向后扫描m列仍然为空，则该行的最大列数为c。
如果c < max_cols，则继续下一行扫描;
否则，更新最大列数(max_cols=c)，并从首行开始，补齐不足max_col的数据。
2. 当某一行r的数据全部都是空，如果再向后扫描m行仍然为空，则最大行数为r。
"""
class ValidCellsParser(object):
	def __init__(self, worksheet, ahead_count = 1, filename = ""):
		super(ValidCellsParser, self).__init__()

		self.worksheet = worksheet
		self.max_rows = 0
		self.max_cols = 0
		self.sheet = []
		self.ahead_count = ahead_count
		self.filename = filename

	def parse(self):
		for i in range(65535):
			if not self.parse_row(self.max_rows):
				break
		else:
			log_error("表格行数过多: %d, %s", self.worksheet.max_column, self.filename)

	def parse_row(self, row):
		cells = self.parse_row_cells(row)
		cells_cache = [cells, ]

		# 遇到空行时，向后检查若干行
		if self.is_all_cells_empty(cells):
			for i in range(self.ahead_count):
				cells = self.parse_row_cells(row + i + 1)
				cells_cache.append(cells)
				if not self.is_all_cells_empty(cells):
					break
			else:
				# 全部都是空行
				return False

		new_max_cols = 0

		for cells in cells_cache:
			self.sheet.append(cells)
			self.max_rows += 1
			new_max_cols = max(new_max_cols, len(cells))

		if new_max_cols > self.max_cols:
			self.max_cols = new_max_cols
			self.backfill()

		return True

	# 注意：worksheet的行列索引是从"1"开始的
	def get_cell(self, r, c):
		return self.worksheet.cell(r + 1, c + 1)

	# 扫描一行数据
	def parse_row_cells(self, row):
		cells = []
		current_max_cols = self.max_cols

		for col in range(65535):
			cell = self.get_cell(row, col)
			count = col + 1
			if cell.value is None:
				if count > current_max_cols + self.ahead_count:
					break
			else:
				if count > current_max_cols:
					current_max_cols = count

			cells.append(cell)
		else:
			log_error("表格列数过多: %d, %s", self.worksheet.max_column, self.filename)

		# 移除超前扫描出的空数据
		while len(cells) > current_max_cols:
			cells.pop()

		return cells

	# 回填解析过的数据
	def backfill(self):
		for row, cells in enumerate(self.sheet):
			cols = len(cells)
			for col in range(cols, self.max_cols):
				cells.append(self.get_cell(row, col))
		return

	def is_all_cells_empty(self, cells):
		for cell in cells:
			if cell.value is not None:
				return False
		return True
