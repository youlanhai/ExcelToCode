# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill
from tps.tp0 import type2string

header_fill = PatternFill(patternType="solid", fgColor="FABF8F")
type_fill = PatternFill(patternType="solid", fgColor="82C3D4")

def create_header(input_file, converter, sheet_index):
	header_row = 2 + 1
	type_row = 3 + 1

	book = openpyxl.load_workbook(input_file)
	table = book.worksheets[sheet_index]

	headers = {}
	for cell in table[header_row]:
		field = cell.value
		if field is None or field == "": break
		headers[field] = True

	hidden_fileds = getattr(converter, "HIDDEN_FIELDS", ())

	new_headers = []
	for info in converter.CONFIG:
		header = info[0].decode("utf-8")
		field = info[1]
		type = info[2]
		if header in headers: continue
		if field in hidden_fileds: continue

		new_headers.append((header, type))

	if len(new_headers) == 0: return True

	new_headers.sort(key = lambda x: x[0])
	col = len(headers) + 1
	for field, type in new_headers:
		cell = table.cell(row = header_row, column = col)
		cell.value = field
		cell.fill = header_fill

		cell = table.cell(row = type_row, column = col)
		cell.value = type2string(type)
		cell.fill = type_fill

		col += 1

	book.save(input_file)
	print "生成表头：", input_file
	return True
